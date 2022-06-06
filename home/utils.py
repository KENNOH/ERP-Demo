import imp
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from django.contrib import messages
import requests
import json
import csv 
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import BLACK, BLUE



def send_confirmation_email(mail_data):
    try:
        subject = mail_data.get('subject')
        email_address = mail_data.get('email')
        message = mail_data.get('message')
        message_template = render_to_string('home/customer_email_template.html',{'message':message})
        email = EmailMessage(subject,message_template,from_email=settings.DEFAULT_FROM_EMAIL,to=[email_address])
        email.content_subtype = "html"
        email.send()
        return "Email sent to customer successfully"
    except:
        return "There was an error sending the email to customer"



def generate_jwt_token(request):
    """username and password vary, we have just hardcoded for easy learning but you can update yours or modify it to accept them as parameters"""
    username = ''
    password = ''

    auth_base_url = 'http://127.0.0.1:8000/'
    url  = '{}{}'.format(auth_base_url,'fetch-auth-token/')
    payload = {'username':username,'password':password}
    headers= {
    'Accept': 'application/json',                         
    'Content-Type': 'application/json; charset=utf-8'
    }
    response = requests.post(url,headers=headers,data=json.dumps(payload))
    if response.status_code == 200:
        data = response.json()['access']
        return data
    else:
        return messages.add_message(request, messages.ERROR, "An error occured fetching the token")
        



def generate_excel_csv(searched_queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = "attached; filename=customer_names.csv"
    customer_names = csv.writer(response)
    customer_names.writerow(['Number','Name','Email','Balance','Gender','Occupation','Phone Number','Created On'])
    for customer in searched_queryset:
        customer_names.writerow([customer.id,customer.name,customer.email,customer.balance,customer.gender,
        customer.occupation,customer.phone_number,customer.created_on])
    return response



def generate_excel_xlsx(searched_queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    """define the file name downloaded"""
    response['Content-Disposition'] = 'attachment; filename=customer_name.xlsx'
    workbook = Workbook()
    workbook.remove(workbook.active)
    """define the font used"""
    header_font = Font(name='Calibre',bold=True)
    centered_allignment = Alignment(horizontal='center')
    """define borders and colors"""
    borders = Border(
        bottom=Side(border_style='medium',color='FF000000'),
        top = Side(border_style='medium',color='FF000000'),
        left = Side(border_style='medium',color='FF000000'),
        right = Side(border_style='medium',color='FF000000'),)
    wrapped_alignment = Alignment(vertical='top',wrap_text=True)
    """add column headers and columns sizes"""
    columns = [
        ('Number',10),
        ('Name',35),
        ('Email',25),
        ('Balance',15),
        ('Gender',15),
        ('Occupation',25),
        ('Phone Number',20),
        ('Created On',25)
    ]
    worksheet = workbook.create_sheet(
        title='Exported Customers',
        index=1
    )
    """add header colors and fill"""
    fill = PatternFill(
        start_color='3366ff',
        end_color='3366ff',
        fill_type='solid'
    )
    row_num = 1
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = borders
        cell.alignment = centered_allignment
        cell.fill = fill
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
    for customer in searched_queryset:
        row_num+=1
        customer_number = customer.id
        customer_name = customer.name
        customer_email = customer.email
        customer_balance = customer.balance
        customer_gender = customer.gender
        customer_occupation = customer.occupation
        customer_phone_number = customer.phone_number
        customer_created_on = str(customer.created_on)
        """define cells and cell types allowed"""
        row = [
            (customer_number,'Normal'),
            (customer_name,'Normal'),
            (customer_email,'Normal'),
            (customer_balance,'Normal'),
            (customer_gender,'Normal'),
            (customer_occupation,'Normal'),
            (customer_phone_number,'Normal'),
            (customer_created_on,'Normal')
        ]
        for col_num, (cell_value, cell_format) in enumerate(row,1):
            cell = worksheet.cell(row=row_num,column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            cell.alignment = wrapped_alignment
    worksheet.freeze_panes = worksheet['A2']
    worksheet.sheet_properties.tabcolor = '00666699'
    workbook.save(response)
    return response
    



    




